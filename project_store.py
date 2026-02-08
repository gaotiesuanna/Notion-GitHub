#!/usr/bin/env python3
"""
项目配置存储层: 统一读写 projects.xlsx
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]

DEFAULT_CONFIG_FILENAME = "projects.xlsx"
LEGACY_JSON_FILENAME = "projects.json"

CATEGORY_HEADERS = ["id", "name", "icon", "order"]
PROJECT_HEADERS = [
    "category_id",
    "id",
    "name",
    "description",
    "github",
    "topics",
    "notion_page_id",
    "order",
]


def _slugify(text: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", (text or "").strip().lower())
    normalized = normalized.strip("-")
    return normalized or "category"


def _ensure_openpyxl():
    if Workbook is None or load_workbook is None:
        raise RuntimeError(
            "缺少依赖 openpyxl。请先安装: pip install -r requirements.txt"
        )


def _resolve_path(config_file: str, base_dir: Path) -> Path:
    path = Path(config_file)
    if not path.is_absolute():
        path = base_dir / path
    return path


def _parse_topics(raw: Any) -> List[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    # 兼容历史上可能存成 JSON 数组字符串
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass
    return [item.strip() for item in text.split(",") if item.strip()]


def _load_from_json(json_path: Path) -> Dict[str, Any]:
    with json_path.open("r", encoding="utf-8") as f:
        config = json.load(f)
    if "categories" in config and isinstance(config["categories"], list):
        return config
    return {"categories": [], "projects": config.get("projects", [])}


def _load_from_xlsx(xlsx_path: Path) -> Dict[str, Any]:
    _ensure_openpyxl()
    wb = load_workbook(xlsx_path, data_only=True)
    categories_ws = wb["categories"] if "categories" in wb.sheetnames else wb.create_sheet("categories")
    projects_ws = wb["projects"] if "projects" in wb.sheetnames else wb.create_sheet("projects")

    categories: List[Dict[str, Any]] = []
    by_category_id: Dict[str, Dict[str, Any]] = {}

    category_rows: List[Tuple[int, Dict[str, Any]]] = []
    for row in categories_ws.iter_rows(min_row=2, values_only=True):
        category_id, name, icon, order = row[:4]
        if not category_id and not name:
            continue
        cid = str(category_id or "").strip() or _slugify(str(name or ""))
        cname = str(name or "").strip() or cid
        category = {
            "id": cid,
            "name": cname,
            "icon": str(icon or "📁"),
            "projects": [],
        }
        by_category_id[cid] = category
        sort_order = int(order) if isinstance(order, int) else 999999
        category_rows.append((sort_order, category))

    category_rows.sort(key=lambda x: (x[0], str(x[1]["name"])))
    categories = [x[1] for x in category_rows]

    project_rows: List[Tuple[int, str, Dict[str, Any]]] = []
    for row in projects_ws.iter_rows(min_row=2, values_only=True):
        category_id, pid, name, description, github, topics, notion_page_id, order = row[:8]
        if not pid and not github:
            continue
        cid = str(category_id or "").strip()
        project = {
            "id": str(pid or "").strip(),
            "name": str(name or "").strip(),
            "description": str(description or "").strip(),
            "github": str(github or "").strip(),
            "topics": _parse_topics(topics),
            "notion_page_id": str(notion_page_id or "").strip(),
        }
        if not project["id"] and project["github"]:
            project["id"] = project["github"].rstrip("/").split("/")[-1].lower()

        sort_order = int(order) if isinstance(order, int) else 999999
        project_rows.append((sort_order, cid, project))

    project_rows.sort(key=lambda x: (x[0], str(x[2].get("id", ""))))
    for _, cid, project in project_rows:
        if not cid:
            cid = "uncategorized"
        category = by_category_id.get(cid)
        if not category:
            category = {"id": cid, "name": cid, "icon": "📁", "projects": []}
            by_category_id[cid] = category
            categories.append(category)
        category["projects"].append(project)

    return {"categories": categories}


def save_projects_config_file(config: Dict[str, Any], config_file: str, base_dir: Path) -> Path:
    _ensure_openpyxl()
    path = _resolve_path(config_file, base_dir)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    categories_ws = wb.create_sheet("categories")
    categories_ws.append(CATEGORY_HEADERS)

    projects_ws = wb.create_sheet("projects")
    projects_ws.append(PROJECT_HEADERS)

    categories = config.get("categories", [])
    for c_idx, category in enumerate(categories):
        cid = str(category.get("id") or _slugify(str(category.get("name", ""))))
        cname = str(category.get("name") or cid)
        icon = str(category.get("icon") or "📁")
        categories_ws.append([cid, cname, icon, c_idx])

        for p_idx, project in enumerate(category.get("projects", [])):
            topics = project.get("topics") or []
            topics_str = ", ".join(str(t).strip() for t in topics if str(t).strip())
            projects_ws.append(
                [
                    cid,
                    str(project.get("id", "")),
                    str(project.get("name", "")),
                    str(project.get("description", "")),
                    str(project.get("github", "")),
                    topics_str,
                    str(project.get("notion_page_id", "")),
                    p_idx,
                ]
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def load_projects_config_file(config_file: str, base_dir: Path) -> Tuple[Dict[str, Any], Path, bool]:
    """
    读取配置并返回: (config, 实际路径, 是否发生自动迁移)
    - 优先读取目标路径
    - 若目标为 xlsx 且不存在,会尝试同目录 legacy json 自动迁移
    """
    path = _resolve_path(config_file, base_dir)
    suffix = path.suffix.lower()
    migrated = False

    if suffix == ".xlsx":
        if path.exists():
            return _load_from_xlsx(path), path, migrated

        legacy_path = path.with_name(LEGACY_JSON_FILENAME)
        if legacy_path.exists():
            config = _load_from_json(legacy_path)
            saved_path = save_projects_config_file(config, str(path), base_dir)
            migrated = True
            return config, saved_path, migrated
        return {"categories": []}, path, migrated

    if suffix == ".json":
        if path.exists():
            config = _load_from_json(path)
            xlsx_path = path.with_suffix(".xlsx")
            save_projects_config_file(config, str(xlsx_path), base_dir)
            migrated = True
            return config, xlsx_path, migrated
        return {"categories": []}, path.with_suffix(".xlsx"), migrated

    # 未显式后缀时默认按 xlsx
    default_path = path.with_suffix(".xlsx")
    if default_path.exists():
        return _load_from_xlsx(default_path), default_path, migrated
    return {"categories": []}, default_path, migrated
