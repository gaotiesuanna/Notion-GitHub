#!/usr/bin/env python3
"""
论文同步脚本：data/papers.xlsx <-> Notion

特性：
- 支持 all / create_only / update_only 三种模式
- 支持从 arXiv 链接自动提取 arXiv ID
- 可用 arXiv API 自动补齐标题、作者、年份（默认只补空字段）
- 同步后回写 notion_page_id 到 Excel
"""

from __future__ import annotations

import os
import re
import time
import xml.etree.ElementTree as ET
import html as html_lib
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from requests.exceptions import ProxyError, RequestException

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]


DEFAULT_PAPERS_FILE = "data/papers.xlsx"
SYNC_MODE_ALIASES = {
    "all": "all",
    "full": "all",
    "create_only": "create_only",
    "create": "create_only",
    "update_only": "update_only",
    "update": "update_only",
}

CATEGORY_HEADERS = ["id", "name", "icon", "order"]
PAPER_HEADERS = [
    "category_id",
    "id",
    "title",
    "authors",
    "venue",
    "year",
    "paper_url",
    "pdf_url",
    "code_url",
    "doi",
    "arxiv_id",
    "keywords",
    "status",
    "rating",
    "notes",
    "notion_page_id",
    "order",
]

ARXIV_ID_PATTERN = re.compile(
    r"arxiv\.org/(?:abs|pdf)/([0-9]{4}\.[0-9]{4,5})(?:v\d+)?(?:\.pdf)?",
    re.IGNORECASE,
)


def _ensure_openpyxl() -> None:
    if Workbook is None or load_workbook is None:
        raise RuntimeError("缺少依赖 openpyxl，请先安装 requirements.txt")


def load_local_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def normalize_sync_mode(raw_mode: str) -> str:
    mode = (raw_mode or "all").strip().lower()
    return SYNC_MODE_ALIASES.get(mode, "all")


def parse_bool_env(raw_value: str, default: bool = False) -> bool:
    if raw_value is None:
        return default
    value = str(raw_value).strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    return default


def normalize_notion_id(raw: str) -> str:
    value = (raw or "").strip().strip('"').strip("'")
    if not value:
        return ""
    if "/" in value:
        value = value.rstrip("/").split("/")[-1]
    if "?" in value:
        value = value.split("?", 1)[0]
    value = value.replace("-", "")
    if re.fullmatch(r"[0-9a-fA-F]{32}", value):
        return f"{value[0:8]}-{value[8:12]}-{value[12:16]}-{value[16:20]}-{value[20:]}"
    return (raw or "").strip()


def split_csv_like(raw: Any) -> List[str]:
    text = str(raw or "").strip()
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def slugify(text: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", (text or "").strip().lower()).strip("-")
    return value or "paper"


def parse_arxiv_id(*values: str) -> str:
    for value in values:
        text = (value or "").strip()
        if not text:
            continue
        if re.fullmatch(r"[0-9]{4}\.[0-9]{4,5}(?:v\d+)?", text):
            return text.split("v", 1)[0]
        match = ARXIV_ID_PATTERN.search(text)
        if match:
            return match.group(1)
    return ""


def make_arxiv_urls(arxiv_id: str) -> Tuple[str, str]:
    aid = (arxiv_id or "").strip()
    if not aid:
        return "", ""
    return f"https://arxiv.org/abs/{aid}", f"https://arxiv.org/pdf/{aid}.pdf"


def fetch_arxiv_metadata(arxiv_id: str, timeout: int = 12) -> Optional[Dict[str, Any]]:
    aid = (arxiv_id or "").strip()
    if not aid:
        return None
    urls = [
        f"https://export.arxiv.org/api/query?id_list={aid}",
        f"http://export.arxiv.org/api/query?id_list={aid}",
    ]
    try:
        response_text = ""
        for url in urls:
            response = requests.get(
                url,
                headers={"User-Agent": "notion-github-sync/1.0"},
                timeout=timeout,
            )
            if response.status_code == 200 and response.text.strip():
                response_text = response.text
                break
        if not response_text:
            return None
        root = ET.fromstring(response_text)
        ns = {"atom": "http://www.w3.org/2005/Atom"}
        entry = root.find("atom:entry", ns)
        if entry is None:
            return None
        title = " ".join((entry.findtext("atom:title", default="", namespaces=ns) or "").split())
        authors = [
            " ".join((node.findtext("atom:name", default="", namespaces=ns) or "").split())
            for node in entry.findall("atom:author", ns)
        ]
        authors = [a for a in authors if a]
        published = (entry.findtext("atom:published", default="", namespaces=ns) or "").strip()
        year = None
        if len(published) >= 4 and published[:4].isdigit():
            year = int(published[:4])
        summary = " ".join((entry.findtext("atom:summary", default="", namespaces=ns) or "").split())
        metadata = {
            "title": title,
            "authors": authors,
            "year": year,
            "summary": summary,
        }
        if metadata.get("title"):
            return metadata

        # 兜底：若 Atom 未返回标题，则从 abs 页面 HTML 提取
        abs_url = f"https://arxiv.org/abs/{aid}"
        page = requests.get(
            abs_url,
            headers={"User-Agent": "notion-github-sync/1.0"},
            timeout=timeout,
        )
        if page.status_code != 200:
            return metadata
        html_text = page.text
        title_match = re.search(
            r'<h1[^>]*class="[^"]*title[^"]*"[^>]*>.*?</h1>',
            html_text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if title_match:
            title_html = title_match.group(0)
            title_html = re.sub(r"<span[^>]*>\s*Title:\s*</span>", "", title_html, flags=re.IGNORECASE)
            title_plain = re.sub(r"<[^>]+>", " ", title_html)
            title_plain = " ".join(html_lib.unescape(title_plain).split()).strip()
            if title_plain:
                metadata["title"] = title_plain
        if not metadata.get("authors"):
            author_matches = re.findall(
                r'<meta\s+name="citation_author"\s+content="([^"]+)"',
                html_text,
                flags=re.IGNORECASE,
            )
            metadata["authors"] = [" ".join(html_lib.unescape(a).split()) for a in author_matches if a.strip()]
        if not metadata.get("year"):
            date_match = re.search(
                r'<meta\s+name="citation_date"\s+content="(\d{4})',
                html_text,
                flags=re.IGNORECASE,
            )
            if date_match:
                metadata["year"] = int(date_match.group(1))
        return metadata
    except Exception:
        return None


def ensure_papers_template(path: Path) -> None:
    _ensure_openpyxl()
    if path.exists():
        return
    wb = Workbook()
    categories_ws = wb.active
    categories_ws.title = "categories"
    categories_ws.append(CATEGORY_HEADERS)
    categories_ws.append(["uncategorized", "未分类", "📚", 0])
    papers_ws = wb.create_sheet("papers")
    papers_ws.append(PAPER_HEADERS)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_papers_config(path: Path) -> Dict[str, Any]:
    _ensure_openpyxl()
    ensure_papers_template(path)
    wb = load_workbook(path, data_only=True)
    categories_ws = wb["categories"]
    papers_ws = wb["papers"]

    categories: List[Dict[str, Any]] = []
    by_category_id: Dict[str, Dict[str, Any]] = {}
    for row in categories_ws.iter_rows(min_row=2, values_only=True):
        cid, name, icon, order = row[:4]
        if not cid and not name:
            continue
        category_id = str(cid or "").strip() or slugify(str(name or ""))
        category_name = str(name or "").strip() or category_id
        sort_order = int(order) if isinstance(order, int) else 999999
        category = {
            "id": category_id,
            "name": category_name,
            "icon": str(icon or "📚"),
            "order": sort_order,
            "papers": [],
        }
        categories.append(category)
        by_category_id[category_id] = category

    categories.sort(key=lambda c: (c["order"], c["name"]))
    if not categories:
        default_category = {"id": "uncategorized", "name": "未分类", "icon": "📚", "order": 0, "papers": []}
        categories.append(default_category)
        by_category_id["uncategorized"] = default_category

    project_rows: List[Tuple[int, str, Dict[str, Any]]] = []
    for row in papers_ws.iter_rows(min_row=2, values_only=True):
        (
            category_id,
            pid,
            title,
            authors,
            venue,
            year,
            paper_url,
            pdf_url,
            code_url,
            doi,
            arxiv_id,
            keywords,
            status,
            rating,
            notes,
            notion_page_id,
            order,
        ) = (list(row[:17]) + [None] * 17)[:17]
        if not any([pid, title, paper_url, arxiv_id, doi, notion_page_id]):
            continue
        arxiv = parse_arxiv_id(str(arxiv_id or ""), str(paper_url or ""), str(pdf_url or ""))
        paper = {
            "category_id": str(category_id or "").strip() or "uncategorized",
            "id": str(pid or "").strip(),
            "title": str(title or "").strip(),
            "authors": split_csv_like(authors),
            "venue": str(venue or "").strip(),
            "year": int(year) if isinstance(year, int) else None,
            "paper_url": str(paper_url or "").strip(),
            "pdf_url": str(pdf_url or "").strip(),
            "code_url": str(code_url or "").strip(),
            "doi": str(doi or "").strip(),
            "arxiv_id": arxiv,
            "keywords": split_csv_like(keywords),
            "status": str(status or "").strip(),
            "rating": int(rating) if isinstance(rating, int) else None,
            "notes": str(notes or "").strip(),
            "notion_page_id": normalize_notion_id(str(notion_page_id or "").strip()),
        }
        if not paper["paper_url"] and paper["arxiv_id"]:
            paper["paper_url"], paper["pdf_url"] = make_arxiv_urls(paper["arxiv_id"])
        if not paper["id"]:
            paper["id"] = slugify(paper["arxiv_id"] or paper["doi"] or paper["title"] or paper["paper_url"])
        sort_order = int(order) if isinstance(order, int) else 999999
        project_rows.append((sort_order, paper["category_id"], paper))

    project_rows.sort(key=lambda item: (item[0], item[2]["id"]))
    for _, cid, paper in project_rows:
        category = by_category_id.get(cid)
        if not category:
            category = {"id": cid, "name": cid, "icon": "📚", "order": 999999, "papers": []}
            categories.append(category)
            by_category_id[cid] = category
        category["papers"].append(paper)

    return {"categories": categories}


def save_papers_config(config: Dict[str, Any], path: Path) -> None:
    _ensure_openpyxl()
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    categories_ws = wb.create_sheet("categories")
    categories_ws.append(CATEGORY_HEADERS)
    papers_ws = wb.create_sheet("papers")
    papers_ws.append(PAPER_HEADERS)

    categories = config.get("categories", [])
    for c_idx, category in enumerate(categories):
        cid = str(category.get("id") or "uncategorized")
        categories_ws.append(
            [
                cid,
                str(category.get("name") or cid),
                str(category.get("icon") or "📚"),
                c_idx,
            ]
        )
        for p_idx, paper in enumerate(category.get("papers", [])):
            papers_ws.append(
                [
                    cid,
                    str(paper.get("id") or ""),
                    str(paper.get("title") or ""),
                    ", ".join(paper.get("authors") or []),
                    str(paper.get("venue") or ""),
                    paper.get("year") if isinstance(paper.get("year"), int) else None,
                    str(paper.get("paper_url") or ""),
                    str(paper.get("pdf_url") or ""),
                    str(paper.get("code_url") or ""),
                    str(paper.get("doi") or ""),
                    str(paper.get("arxiv_id") or ""),
                    ", ".join(paper.get("keywords") or []),
                    str(paper.get("status") or ""),
                    paper.get("rating") if isinstance(paper.get("rating"), int) else None,
                    str(paper.get("notes") or ""),
                    str(paper.get("notion_page_id") or ""),
                    p_idx,
                ]
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


class PaperNotionSync:
    def __init__(self, notion_token: str, database_id: str, force_arxiv_title: bool = False):
        self.database_id = database_id
        self.force_arxiv_title = force_arxiv_title
        self.notion_headers = {
            "Authorization": f"Bearer {notion_token}",
            "Content-Type": "application/json",
            "Notion-Version": "2022-06-28",
        }
        self.notion_session = requests.Session()
        self.notion_direct_session = requests.Session()
        self.notion_direct_session.trust_env = False
        self._database_properties: Optional[Dict[str, Any]] = None

    def notion_request(self, method: str, url: str, **kwargs):
        timeout = kwargs.pop("timeout", 12)
        try:
            return self.notion_session.request(method, url, headers=self.notion_headers, timeout=timeout, **kwargs)
        except ProxyError:
            try:
                return self.notion_direct_session.request(
                    method,
                    url,
                    headers=self.notion_headers,
                    timeout=timeout,
                    **kwargs,
                )
            except RequestException as e:
                print(f"  ⚠ Notion 请求失败(直连): {e}")
                return None
        except RequestException as e:
            print(f"  ⚠ Notion 请求失败: {e}")
            return None

    def get_database_properties(self) -> Dict[str, Any]:
        if self._database_properties is not None:
            return self._database_properties
        url = f"https://api.notion.com/v1/databases/{self.database_id}"
        response = self.notion_request("GET", url)
        if response is None:
            self._database_properties = {}
        elif response.status_code == 200:
            self._database_properties = response.json().get("properties", {})
        else:
            self._database_properties = {}
        return self._database_properties

    def get_property_type(self, property_name: str) -> str:
        return self.get_database_properties().get(property_name, {}).get("type", "")

    def _set_text_property(self, properties: Dict[str, Any], name: str, content: str, preferred: str) -> None:
        text = (content or "").strip()
        if not text:
            return
        ptype = self.get_property_type(name) or preferred
        if ptype == "title":
            properties[name] = {"title": [{"text": {"content": text[:2000]}}]}
        elif ptype == "rich_text":
            properties[name] = {"rich_text": [{"text": {"content": text[:2000]}}]}

    def _set_url_property(self, properties: Dict[str, Any], name: str, url_value: str) -> None:
        value = (url_value or "").strip()
        if not value:
            return
        if (self.get_property_type(name) or "url") == "url":
            properties[name] = {"url": value}

    def _set_number_property(self, properties: Dict[str, Any], name: str, value: Any) -> None:
        if value is None:
            return
        if (self.get_property_type(name) or "number") == "number":
            properties[name] = {"number": value}

    def _set_select_property(self, properties: Dict[str, Any], name: str, value: str) -> None:
        text = (value or "").strip()
        if not text:
            return
        ptype = self.get_property_type(name) or "select"
        if ptype == "select":
            properties[name] = {"select": {"name": text}}
        elif ptype == "rich_text":
            properties[name] = {"rich_text": [{"text": {"content": text[:2000]}}]}

    def _set_multi_select_property(self, properties: Dict[str, Any], name: str, values: List[str]) -> None:
        cleaned = [v.strip() for v in values if v and v.strip()]
        if not cleaned:
            return
        ptype = self.get_property_type(name) or "multi_select"
        if ptype == "multi_select":
            properties[name] = {"multi_select": [{"name": v[:100]} for v in cleaned[:20]]}
        elif ptype == "rich_text":
            properties[name] = {"rich_text": [{"text": {"content": ", ".join(cleaned)[:2000]}}]}

    def build_notion_properties(self, paper: Dict[str, Any], category_name: str) -> Dict[str, Any]:
        properties: Dict[str, Any] = {}
        self._set_text_property(properties, "标题", paper.get("title", ""), "title")
        self._set_url_property(properties, "论文链接", paper.get("paper_url", ""))
        self._set_url_property(properties, "PDF链接", paper.get("pdf_url", ""))
        self._set_text_property(properties, "arXiv ID", paper.get("arxiv_id", ""), "rich_text")
        self._set_text_property(properties, "作者", ", ".join(paper.get("authors") or []), "rich_text")
        self._set_number_property(properties, "年份", paper.get("year"))
        self._set_select_property(properties, "会议/期刊", paper.get("venue", ""))
        self._set_multi_select_property(properties, "关键词", paper.get("keywords") or [])
        self._set_select_property(properties, "状态", paper.get("status", ""))
        self._set_number_property(properties, "评分", paper.get("rating"))
        self._set_text_property(properties, "笔记", paper.get("notes", ""), "rich_text")
        self._set_select_property(properties, "分类", category_name)
        self._set_text_property(properties, "DOI", paper.get("doi", ""), "rich_text")
        self._set_url_property(properties, "Code链接", paper.get("code_url", ""))
        return properties

    def query_database(self, query_body: Dict[str, Any]) -> List[Dict[str, Any]]:
        url = f"https://api.notion.com/v1/databases/{self.database_id}/query"
        response = self.notion_request("POST", url, json=query_body)
        if response is None:
            return []
        if response.status_code != 200:
            return []
        return response.json().get("results", [])

    def find_existing_page_id(self, paper: Dict[str, Any]) -> Optional[str]:
        arxiv_id = (paper.get("arxiv_id") or "").strip()
        doi = (paper.get("doi") or "").strip()
        paper_url = (paper.get("paper_url") or "").strip()

        if arxiv_id and self.get_property_type("arXiv ID") in {"rich_text", "title"}:
            results = self.query_database(
                {"filter": {"property": "arXiv ID", "rich_text": {"equals": arxiv_id}}, "page_size": 5}
            )
            if results:
                return normalize_notion_id(results[0].get("id", ""))
        if doi and self.get_property_type("DOI") in {"rich_text", "title"}:
            results = self.query_database({"filter": {"property": "DOI", "rich_text": {"equals": doi}}, "page_size": 5})
            if results:
                return normalize_notion_id(results[0].get("id", ""))
        if paper_url and self.get_property_type("论文链接") == "url":
            results = self.query_database({"filter": {"property": "论文链接", "url": {"equals": paper_url}}, "page_size": 5})
            if results:
                return normalize_notion_id(results[0].get("id", ""))
        return None

    def create_page(self, properties: Dict[str, Any]) -> Optional[str]:
        url = "https://api.notion.com/v1/pages"
        body = {"parent": {"database_id": self.database_id}, "properties": properties}
        response = self.notion_request("POST", url, json=body)
        if response is None:
            print("  ✗ 创建失败: 请求异常")
            return None
        if response.status_code == 200:
            return normalize_notion_id(response.json().get("id", ""))
        print(f"  ✗ 创建失败({response.status_code}): {response.text}")
        return None

    def update_page(self, page_id: str, properties: Dict[str, Any]) -> str:
        url = f"https://api.notion.com/v1/pages/{page_id}"
        response = self.notion_request("PATCH", url, json={"properties": properties})
        if response is None:
            return "error"
        if response.status_code == 200:
            return "ok"
        if response.status_code == 404:
            return "not_found"
        print(f"  ✗ 更新失败({response.status_code}): {response.text}")
        return "error"

    def sync_one(self, paper: Dict[str, Any], category_name: str) -> Tuple[Optional[str], str]:
        arxiv_id = parse_arxiv_id(paper.get("arxiv_id", ""), paper.get("paper_url", ""), paper.get("pdf_url", ""))
        if arxiv_id:
            paper["arxiv_id"] = arxiv_id
            if not paper.get("paper_url"):
                paper["paper_url"], _ = make_arxiv_urls(arxiv_id)
            if not paper.get("pdf_url"):
                _, paper["pdf_url"] = make_arxiv_urls(arxiv_id)

        if arxiv_id:
            metadata = fetch_arxiv_metadata(arxiv_id)
            if metadata:
                # FORCE_ARXIV_TITLE=true 时覆盖现有标题；否则只补空标题。
                if metadata.get("title") and (self.force_arxiv_title or not paper.get("title")):
                    paper["title"] = metadata["title"]
                if not paper.get("authors") and metadata.get("authors"):
                    paper["authors"] = metadata["authors"]
                if not paper.get("year") and metadata.get("year"):
                    paper["year"] = metadata["year"]

        if not paper.get("title"):
            paper["title"] = paper.get("doi") or paper.get("id") or "Untitled Paper"

        properties = self.build_notion_properties(paper, category_name)
        page_id = normalize_notion_id(paper.get("notion_page_id", ""))
        if page_id:
            status = self.update_page(page_id, properties)
            if status == "ok":
                return page_id, "updated"
            if status == "error":
                return None, "failed"

        recovered_id = self.find_existing_page_id(paper)
        if recovered_id:
            status = self.update_page(recovered_id, properties)
            if status == "ok":
                return recovered_id, "updated"

        created_id = self.create_page(properties)
        if created_id:
            return created_id, "created"
        return None, "failed"


def flatten_papers(config: Dict[str, Any]) -> List[Tuple[Dict[str, Any], str]]:
    rows: List[Tuple[Dict[str, Any], str]] = []
    for category in config.get("categories", []):
        category_name = str(category.get("name") or "")
        for paper in category.get("papers", []):
            rows.append((paper, category_name))
    return rows


def main() -> None:
    project_root = Path(__file__).resolve().parent.parent
    env_file = project_root / ".env"
    if env_file.exists():
        if load_dotenv:
            load_dotenv(dotenv_path=env_file)
        else:
            load_local_env_file(env_file)

    notion_token = os.environ.get("NOTION_TOKEN", "").strip()
    database_id = normalize_notion_id(os.environ.get("NOTION_PAPERS_DATABASE_ID", ""))
    papers_file = os.environ.get("PAPERS_FILE", DEFAULT_PAPERS_FILE).strip() or DEFAULT_PAPERS_FILE
    sync_mode = normalize_sync_mode(os.environ.get("SYNC_MODE", "all"))
    force_arxiv_title = parse_bool_env(os.environ.get("FORCE_ARXIV_TITLE", "false"), default=False)

    if not notion_token:
        print("❌ 未设置 NOTION_TOKEN")
        return
    if not database_id:
        print("❌ 未设置 NOTION_PAPERS_DATABASE_ID")
        return
    if not re.fullmatch(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", database_id):
        print("❌ NOTION_PAPERS_DATABASE_ID 格式不合法（应为数据库 ID 或可解析出数据库 ID 的 URL）")
        return

    config_path = project_root / papers_file if not Path(papers_file).is_absolute() else Path(papers_file)
    config = load_papers_config(config_path)
    papers_with_category = flatten_papers(config)
    if not papers_with_category:
        print(f"⚠ 没有论文记录，请先编辑 {config_path}")
        return

    syncer = PaperNotionSync(
        notion_token=notion_token,
        database_id=database_id,
        force_arxiv_title=force_arxiv_title,
    )
    print(f"开始同步论文: {len(papers_with_category)} 条，模式: {sync_mode}")
    print(f"标题覆盖: {'开启' if force_arxiv_title else '关闭'} (FORCE_ARXIV_TITLE)")

    created = 0
    updated = 0
    skipped = 0
    failed = 0

    for idx, (paper, category_name) in enumerate(papers_with_category, 1):
        has_page_id = bool((paper.get("notion_page_id") or "").strip())
        if sync_mode == "create_only" and has_page_id:
            skipped += 1
            continue
        if sync_mode == "update_only" and not has_page_id:
            skipped += 1
            continue

        print(f"\n[{idx}/{len(papers_with_category)}] {paper.get('id') or 'paper'}")
        page_id, action = syncer.sync_one(paper, category_name)
        if page_id:
            paper["notion_page_id"] = page_id
            if action == "created":
                created += 1
            else:
                updated += 1
        else:
            failed += 1
        if idx < len(papers_with_category):
            time.sleep(0.3)

    save_papers_config(config, config_path)
    print("\n同步完成")
    print(f"  ✓ 创建: {created}")
    print(f"  ✓ 更新: {updated}")
    print(f"  - 跳过: {skipped}")
    if failed:
        print(f"  ✗ 失败: {failed}")
    print(f"已回写: {config_path}")


if __name__ == "__main__":
    main()
